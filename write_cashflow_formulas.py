"""
כתיבת גיליונות תזרים מבוססי-נוסחאות (ניתנים למעקב מלא)
======================================================
כל תא הוא נוסחה: העלויות נמשכות מגיליון "מאוחד", הפריסה מחושבת מול
תא ההנחות, ובלוק המימון (יתרה, ריבית) הוא שרשרת נוסחאות. אין מספרים קשיחים.
"""
import sys, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

FONT = "Arial"
HDR = PatternFill("solid", fgColor="4F2D7F")
SUB = PatternFill("solid", fgColor="EDE7F6")
TOT = PatternFill("solid", fgColor="FFE0B2")
FIN = PatternFill("solid", fgColor="FFF3E0")
INP = PatternFill("solid", fgColor="FFF9C4")   # צהוב לקלט
WHT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
BLD = Font(name=FONT, bold=True, size=10)
REG = Font(name=FONT, size=10)
BLUE = Font(name=FONT, size=10, color="0000FF")   # קלט (מוסכמה: כחול)
CUR = '#,##0;(#,##0);-'
PCT = '0.00%'
CEN = Alignment(horizontal="center")
RGT = Alignment(horizontal="right")

# ---- מיפוי שורות מגיליון מאוחד ----
M = "מאוחד"
def ref(cell): return f"={M}!{cell}"


def build(path_in, path_out, overrides=None):
    from openpyxl import Workbook
    ov = overrides or {}
    def OV(k, d): return ov[k] if ov.get(k) is not None else d
    # קוראים את ערכי "מאוחד" (מוקפאים) — כדי להימנע מתלות בקישור חיצוני שבור
    src_m = load_workbook(path_in, data_only=True)["מאוחד"]
    wb = Workbook()
    wb.remove(wb.active)
    msheet = wb.create_sheet(M)
    msheet.sheet_view.rightToLeft = True
    for row in src_m.iter_rows():
        for c in row:
            if c.value is not None and not (isinstance(c.value, str) and c.value.startswith("=")):
                msheet.cell(row=c.row, column=c.column, value=c.value)
    msheet.column_dimensions["B"].width = 40

    # =================== גיליון תזרים הכנסות ===================
    R = wb.create_sheet("תזרים הכנסות", 0)
    R.sheet_view.rightToLeft = True
    def rset(r, c, v, font=REG, fmt=None, fill=None, al=None):
        cell = R.cell(row=r, column=c, value=v)
        cell.font = font
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        if al: cell.alignment = al
        return cell

    rset(1, 2, "תזרים הכנסות (פדיון)", Font(name=FONT, bold=True, size=13, color="4F2D7F"))
    # הנחות (קלט/נגזר)
    A = {}
    rows_assum = [
        ("שווי פדיון (ללא מע\"מ)", ref("J151"), None),
        ("חודשי בנייה", OV("months", 42), INP),
        ("תקופת מבצע (חודשים)", OV("promo", 1), INP),
        ("תקבול ראשון (מקדמה)", OV("down", 0.20), INP),
        ("חלק מסלול 20%", OV("track20", 0.20), INP),
    ]
    r0 = 3
    keys = ["pidyon", "months", "promo", "down", "track20"]
    for i, (lbl, val, fill) in enumerate(rows_assum):
        rset(r0+i, 2, lbl, REG, al=RGT)
        c = rset(r0+i, 3, val, BLUE if fill else REG,
                 fmt=(CUR if isinstance(val, str) or val > 1 else PCT), fill=fill)
        A[keys[i]] = f"$C${r0+i}"
    # נגזרים
    der_r = r0 + len(rows_assum) + 1
    ders = [
        ("מסלול 20% (סכום)", f"={A['pidyon']}*{A['track20']}"),
        ("מסלול רגיל (סכום)", f"={A['pidyon']}*(1-{A['track20']})"),
        ("מס' אצוות", f"={A['months']}-{A['promo']}"),
    ]
    D = {}
    dkeys = ["I4", "I5", "nbatch"]
    for i, (lbl, f) in enumerate(ders):
        rset(der_r+i, 2, lbl, REG, al=RGT)
        rset(der_r+i, 3, f, REG, fmt=CUR)
        D[dkeys[i]] = f"$C${der_r+i}"
    batchval = f"({D['I5']}/{D['nbatch']})"
    batchdown = f"({batchval}*{A['down']})"
    batchbal = f"({batchval}*(1-{A['down']}))"

    MONTHS = int(OV("months", 42))
    n = MONTHS + 1
    MC0 = 4                       # עמודת חודש 0
    def mcol(m): return MC0 + m
    hdr_r = der_r + len(ders) + 2
    midx_r = hdr_r + 1
    rset(hdr_r, 2, "חודש", WHT, fill=HDR)
    rset(midx_r, 2, "אינדקס", REG, al=RGT)
    for m in range(n):
        rset(hdr_r, mcol(m), "היתר" if m == 0 else m, WHT, fill=HDR, al=CEN)
        rset(midx_r, mcol(m), m, REG, al=CEN)
    tot_col = mcol(MONTHS) + 1
    rset(hdr_r, tot_col, "סה\"כ", WHT, fill=HDR)

    # אצוות
    batch_r0 = midx_r + 1
    # מסלול 20%
    rset(batch_r0, 2, "מסלול 20% (חתימה+בלון)", REG, al=RGT)
    for m in range(n):
        mi = f"{CL(mcol(m))}${midx_r}"
        f = (f"=IF({mi}={A['promo']},{D['I4']}*{A['down']},"
             f"IF({mi}={A['months']},{D['I4']}*(1-{A['down']}),0))")
        rset(batch_r0, mcol(m), f, REG, fmt=CUR)
    rset(batch_r0, tot_col, f"=SUM({CL(MC0)}{batch_r0}:{CL(mcol(MONTHS))}{batch_r0})", REG, fmt=CUR)

    # אצוות רגילות: נמכרות בחודשים promo+1..M
    br = batch_r0 + 1
    first_reg = br
    for k in range(2, MONTHS + 1):          # חודש מכירה k (promo=1)
        rset(br, 2, f"אצווה חודש {k}", REG, al=RGT)
        for m in range(n):
            col = mcol(m)
            if m == k:
                f = f"={batchdown}"
            elif m > k:
                f = f"={batchbal}/({A['months']}-{k})"
            else:
                f = 0
            rset(br, col, f, REG, fmt=CUR)
        # אצווה אחרונה (k=M): מקדמה+יתרה מלאה
        if k == MONTHS:
            rset(br, mcol(k), f"={batchval}", REG, fmt=CUR)
        rset(br, tot_col, f"=SUM({CL(MC0)}{br}:{CL(mcol(MONTHS))}{br})", REG, fmt=CUR)
        br += 1
    last_reg = br - 1

    # שורת סה"כ חודשי
    rev_r = br + 1
    rset(rev_r, 2, "הכנסות לחודש", BLD, fill=SUB)
    for m in range(n):
        col = mcol(m)
        f = f"=SUM({CL(col)}{batch_r0}:{CL(col)}{last_reg})"
        rset(rev_r, col, f, BLD, fmt=CUR, fill=SUB)
    rset(rev_r, tot_col, f"=SUM({CL(MC0)}{rev_r}:{CL(mcol(MONTHS))}{rev_r})", BLD, fmt=CUR, fill=SUB)

    R.column_dimensions["B"].width = 20
    for m in range(n + 1):
        R.column_dimensions[CL(MC0 + m)].width = 12
    REV_SHEET = "תזרים הכנסות"
    REV_ROW = rev_r

    # =================== גיליון תזרים הוצאות ===================
    E = wb.create_sheet("תזרים הוצאות", 0)
    E.sheet_view.rightToLeft = True
    def eset(r, c, v, font=REG, fmt=None, fill=None, al=None):
        cell = E.cell(row=r, column=c, value=v)
        cell.font = font
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        if al: cell.alignment = al
        return cell

    eset(1, 2, "תזרים הוצאות ומימון", Font(name=FONT, bold=True, size=13, color="4F2D7F"))

    # ---- הנחות עבודה ----
    P = {}
    assum = [
        ("months", "חודשי בנייה", OV("months", 42), INP, "0"),
        ("rate", "ריבית שנתית", OV("rate", 0.055), INP, PCT),
        ("equity", "הון עצמי (שיעור)", OV("equity", 0.25), INP, PCT),
        ("permit_share", "תכנון — חלק בהיתר", OV("permit_share", 0.50), INP, PCT),
        ("fee_acc", "עמלת ליווי", OV("fee_acc", 0.005), INP, PCT),
        ("fee_sale", "ערבות חוק מכר", OV("fee_sale", 0.008), INP, PCT),
        ("fee_rent", "ערבות שכ\"ד", OV("fee_rent", 0.035), INP, PCT),
        ("fee_own", "ערבות בעלים", OV("fee_own", 0.007), INP, PCT),
        ("fee_nu", "עמלת אי-ניצול", OV("fee_nu", 0.004), INP, PCT),
        ("bs", "מרתף — חודש התחלה", OV("bs", 2), INP, "0"),
        ("bd", "מרתף — משך", OV("bd", 9), INP, "0"),
        ("ss", "שלד — חודש התחלה", OV("ss", 10), INP, "0"),
        ("sd", "שלד — משך", OV("sd", 12), INP, "0"),
        ("fs", "גמר — חודש התחלה", OV("fs", 20), INP, "0"),
        ("fd", "גמר — משך", OV("fd", 23), INP, "0"),
    ]
    ar = 3
    for i, (k, lbl, val, fill, fmt) in enumerate(assum):
        eset(ar+i, 2, lbl, REG, al=RGT)
        eset(ar+i, 3, val, BLUE, fmt=fmt, fill=fill)
        P[k] = f"$C${ar+i}"
    # ערכי בסיס מ"מאוחד"
    base_r = ar + len(assum) + 1
    bvals = [
        ("owners_full", "שווי דירות בעלים (כולל מרפסות)",
         f"={M}!I139*{M}!I88+{M}!I139*({M}!J88+{M}!K88)*0.5"),
        ("owners_main", "שווי דירות בעלים (עיקרי)", f"={M}!I139*{M}!I88"),
        ("rent_total", "שכ\"ד דיור חלוף (סה\"כ)", ref("G194")),
        ("total_cost", "סה\"כ עלות הקמה", ref("G213")),
    ]
    for i, (k, lbl, f) in enumerate(bvals):
        eset(base_r+i, 2, lbl, REG, al=RGT)
        eset(base_r+i, 3, f, REG, fmt=CUR)
        P[k] = f"$C${base_r+i}"

    # ---- טבלת חודשים ----
    hr = base_r + len(bvals) + 2
    mi_r = hr + 1
    eset(hr, 2, "סעיף", WHT, fill=HDR)
    eset(mi_r, 2, "אינדקס", REG, al=RGT)
    for m in range(n):
        eset(hr, mcol(m), "היתר" if m == 0 else m, WHT, fill=HDR, al=CEN)
        eset(mi_r, mcol(m), m, REG, al=CEN)
    eset(hr, tot_col, "סה\"כ", WHT, fill=HDR)

    def mi(m): return f"{CL(mcol(m))}${mi_r}"

    # הגדרת שורות עלות: (label, total_formula, rule)
    cost = [
        ("תכנון ויועצים", ref("G182"), "permit_linear"),
        ("היטל השבחה", ref("G207"), "at_permit"),
        ("מס רכישה", ref("G206"), "at_permit"),
        ("אגרות והיטלי פיתוח", f"={M}!G179+{M}!G180+{M}!G181", "at_permit"),
        ("הריסה ופיתוח", f"={M}!G157+{M}!G159", "at_permit"),
        ("חיבור חשמל מגורים", ref("G183"), "month1"),
        ("חיבור חשמל מסחרי", ref("G184"), "month1"),
        ("משפטיות", ref("G187"), "linear"),
        ("הובלות", ref("G196"), "start_end"),
        ("יועצי דיירים (עו\"ד/מפקח/שמאי/קרן)", f"={M}!G198+{M}!G199+{M}!G200+{M}!G201", "linear"),
        ("שכ\"ד דיור חלוף", ref("G194"), "rent"),
        ("שיווק ופרסום", ref("G185"), "linear"),
        ("ניהול תקורה ופיקוח", f"={M}!G186+{M}!G188", "linear"),
        ("בלת\"מ", ref("G189"), "linear"),
        ("מרתפים", ref("G170"), "win_bsmt"),
        ("שלד", f"=({M}!G175-{M}!G170)*0.6", "win_skel"),
    ]
    cost_r0 = mi_r + 1
    r = cost_r0
    total_cells = {}     # שם שורה -> תא total (C)
    for lbl, tf, rule in cost:
        eset(r, 2, lbl, REG, al=RGT)
        tcell = eset(r, 3, tf, REG, fmt=CUR)   # סכום מקור מ"מאוחד"
        total_cells[lbl] = f"$C${r}"
        _spread(E, r, rule, P, mi, mcol, n, MONTHS)
        eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR)
        r += 1
    # גמר = סה"כ בנייה ישירה - שלד - מרתף
    eset(r, 2, "גמר", REG, al=RGT)
    skel = total_cells["שלד"]; bsmt = total_cells["מרתפים"]
    eset(r, 3, f"={M}!G175-{skel}-{bsmt}", REG, fmt=CUR)
    total_cells["גמר"] = f"$C${r}"
    _spread(E, r, "win_fin", P, mi, mcol, n, MONTHS)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR)
    cost_rlast = r
    r += 1

    # סה"כ הוצאות
    exp_r = r
    eset(r, 2, "סה\"כ הוצאות", BLD, fill=SUB)
    for m in range(n):
        col = mcol(m)
        eset(r, col, f"=SUM({CL(col)}{cost_r0}:{CL(col)}{cost_rlast})", BLD, fmt=CUR, fill=SUB)
    eset(r, tot_col, f"=SUM({CL(MC0)}{exp_r}:{CL(mcol(MONTHS))}{exp_r})", BLD, fmt=CUR, fill=SUB)
    EXP_TOTAL = f"${CL(tot_col)}${exp_r}"
    r += 2

    # ---- ערבויות ועמלות ----
    eset(r, 2, "ערבויות ועמלות", BLD); r += 1
    g_r0 = r
    # ליווי (חד-פעמי בחודש 1)
    eset(r, 2, "עמלת ליווי", REG, al=RGT)
    for m in range(n):
        eset(r, mcol(m), f"=IF({mi(m)}=1,{EXP_TOTAL}*{P['fee_acc']},0)", REG, fmt=CUR)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR); acc_r = r; r += 1
    # חוק מכר
    eset(r, 2, "ערבות חוק מכר", REG, al=RGT)
    for m in range(n):
        rc = f"'{REV_SHEET}'!{CL(mcol(m))}{REV_ROW}"
        eset(r, mcol(m), f"=IF({mi(m)}>=1,{rc}*{P['fee_sale']}/12,0)", REG, fmt=CUR)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR); sale_r = r; r += 1
    # שכ"ד
    eset(r, 2, "ערבות שכ\"ד", REG, al=RGT)
    for m in range(n):
        eset(r, mcol(m), f"=IF({mi(m)}>=1,{P['fee_rent']}*{P['rent_total']}/{P['months']},0)", REG, fmt=CUR)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR); rent_r = r; r += 1
    # בעלים
    eset(r, 2, "ערבות בעלים", REG, al=RGT)
    for m in range(n):
        f = (f"=IF({mi(m)}=1,{P['fee_own']}/12*{P['owners_full']},"
             f"IF({mi(m)}>=2,{P['fee_own']}/12*{P['owners_main']},0))")
        eset(r, mcol(m), f, REG, fmt=CUR)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR); own_r = r; r += 1
    # אי-ניצול
    eset(r, 2, "עמלת אי-ניצול", REG, al=RGT)
    for m in range(n):
        col = mcol(m)
        cum = f"SUM({CL(MC0)}{exp_r}:{CL(col)}{exp_r})"
        eset(r, col, f"=IF({mi(m)}>=1,{P['fee_nu']}/12*MAX({EXP_TOTAL}-{cum},0),0)", REG, fmt=CUR)
    eset(r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(MONTHS))}{r})", REG, fmt=CUR); nu_r = r
    g_rlast = r; r += 2

    # ---- בלוק מימון ----
    eset(r, 2, "סך הוצאות כולל עמלות", BLD)
    incl_r = r
    for m in range(n):
        col = mcol(m)
        eset(r, col, f"={CL(col)}{exp_r}+SUM({CL(col)}{g_r0}:{CL(col)}{g_rlast})", BLD, fmt=CUR)
    r += 1
    eset(r, 2, "הכנסות", REG)
    inc_rev_r = r
    for m in range(n):
        eset(r, mcol(m), f"='{REV_SHEET}'!{CL(mcol(m))}{REV_ROW}", REG, fmt=CUR)
    r += 1
    eset(r, 2, "תזרים חודשי", REG)
    cf_r = r
    for m in range(n):
        col = mcol(m)
        eset(r, col, f"={CL(col)}{inc_rev_r}-{CL(col)}{incl_r}", REG, fmt=CUR)
    r += 1
    eset(r, 2, "הון עצמי", REG)
    eq_r = r
    cap = f"{P['equity']}*{EXP_TOTAL}"
    for m in range(n):
        col = mcol(m)
        if m == 0:
            f = f"=MAX(MIN({CL(col)}{incl_r},{cap}),0)"
        else:
            prior = f"SUM({CL(MC0)}{eq_r}:{CL(mcol(m-1))}{eq_r})"
            f = f"=MAX(MIN({CL(col)}{incl_r},{cap}-{prior}),0)"
        eset(r, col, f, REG, fmt=CUR)
    r += 1
    eset(r, 2, "יתרה לפני ריבית", REG)
    bb_r = r
    for m in range(n):
        col = mcol(m)
        if m == 0:
            f = f"={CL(col)}{eq_r}+{CL(col)}{cf_r}"
        else:
            f = f"={CL(mcol(m-1))}{bb_r+2}+{CL(col)}{eq_r}+{CL(col)}{cf_r}"  # prev after-interest
        eset(r, col, f, REG, fmt=CUR)
    r += 1
    eset(r, 2, "ריבית", REG)
    int_r = r
    for m in range(n):
        col = mcol(m)
        eset(r, col, f"=IF({CL(col)}{bb_r}<0,{CL(col)}{bb_r}*{P['rate']}/12,0)", REG, fmt=CUR)
    r += 1
    eset(r, 2, "יתרה לסוף חודש", BLD)
    ai_r = r
    for m in range(n):
        col = mcol(m)
        eset(r, col, f"={CL(col)}{bb_r}+{CL(col)}{int_r}", BLD, fmt=CUR)
    r += 3

    # ---- סיכום מימון ----
    eset(r, 2, "סיכום מימון", Font(name=FONT, bold=True, size=12, color="4F2D7F")); r += 1
    fin_rows = [
        ("ריבית נצברת", f"=-SUM({CL(MC0)}{int_r}:{CL(mcol(MONTHS))}{int_r})"),
        ("עמלת ליווי", f"={CL(tot_col)}{acc_r}"),
        ("ערבות חוק מכר", f"={CL(tot_col)}{sale_r}"),
        ("ערבות שכ\"ד", f"={CL(tot_col)}{rent_r}"),
        ("ערבות בעלים", f"={CL(tot_col)}{own_r}"),
        ("עמלת אי-ניצול", f"={CL(tot_col)}{nu_r}"),
    ]
    fr0 = r
    for lbl, f in fin_rows:
        eset(r, 2, lbl, REG, al=RGT)
        eset(r, 3, f, REG, fmt=CUR, fill=FIN); r += 1
    eset(r, 2, "סה\"כ מימון", BLD, al=RGT)
    eset(r, 3, f"=ROUND(SUM(C{fr0}:C{r-1}),-4)", BLD, fmt=CUR, fill=TOT); tot_fin_r = r; r += 1
    eset(r, 2, "סה\"כ עלות הקמה", REG, al=RGT)
    eset(r, 3, f"={P['total_cost']}", REG, fmt=CUR); tc_r = r; r += 1
    eset(r, 2, "אחוז מימון", Font(name=FONT, bold=True, size=12), al=RGT)
    eset(r, 3, f"=C{tot_fin_r}/C{tc_r}", Font(name=FONT, bold=True, size=12, color="4F2D7F"), fmt=PCT, fill=TOT)

    E.column_dimensions["B"].width = 30
    for m in range(n + 1):
        E.column_dimensions[CL(MC0 + m)].width = 12
    E.freeze_panes = f"{CL(MC0)}{hr+2}"

    wb.save(path_out)


def _spread(ws, r, rule, P, mi, mcol, n, MONTHS):
    """כותב נוסחת פריסה בכל תא חודשי לפי הכלל. תא הסכום הוא $C$r."""
    from openpyxl.utils import get_column_letter as CL
    T = f"$C${r}"
    for m in range(n):
        col = mcol(m); mm = mi(m)
        if rule == "at_permit":
            f = f"=IF({mm}=0,{T},0)"
        elif rule == "month1":
            f = f"=IF({mm}=1,{T},0)"
        elif rule == "linear":
            f = f"=IF({mm}>=1,{T}/{P['months']},0)"
        elif rule == "permit_linear":
            f = f"=IF({mm}=0,{T}*{P['permit_share']},IF({mm}>=1,{T}*(1-{P['permit_share']})/{P['months']},0))"
        elif rule == "start_end":
            f = f"=IF({mm}=1,{T}/2,IF({mm}={P['months']},{T}/2,0))"
        elif rule == "rent":
            ann = f"{T}/{P['months']}*12"
            f = (f"=IF(AND({mm}>=1,MOD({mm}-1,12)=0),"
                 f"MIN({ann},MAX({T}-({ann})*INT(({mm}-1)/12),0)),0)")
        elif rule in ("win_bsmt", "win_skel", "win_fin"):
            s = {"win_bsmt": P['bs'], "win_skel": P['ss'], "win_fin": P['fs']}[rule]
            d = {"win_bsmt": P['bd'], "win_skel": P['sd'], "win_fin": P['fd']}[rule]
            f = f"=IF(AND({mm}>={s},{mm}<={s}+{d}-1),{T}/{d},0)"
        else:
            f = 0
        c = ws.cell(row=r, column=col, value=f)
        c.font = REG; c.number_format = CUR


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "new.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "תחשיב_עם_תזרים_נוסחאות.xlsx"
    build(inp, out)
    print("נוצר:", out)
