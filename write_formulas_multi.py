"""
כתיבת קובץ תזרים רב-מתחמי מבוסס-נוסחאות (ניתן לבדיקה)
=====================================================
לכל מתחם: גיליון "תזרים הכנסות מתחם N" (מדרגה עם נוסחאות + שורת SUM)
ו-"תזרים הוצאות מתחם N" (פריסה בנוסחאות + בלוק מימון + אחוז). גיליון "סיכום".
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as CL
import extract as X

FONT = "Arial"
HDR = PatternFill("solid", fgColor="4F2D7F"); SUB = PatternFill("solid", fgColor="EDE7F6")
TOT = PatternFill("solid", fgColor="FFE0B2"); FIN = PatternFill("solid", fgColor="FFF3E0")
INP = PatternFill("solid", fgColor="FFF9C4"); OKF = PatternFill("solid", fgColor="C8E6C9"); BADF = PatternFill("solid", fgColor="FFCDD2")
WHT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
BLD = Font(name=FONT, bold=True, size=10); REG = Font(name=FONT, size=10); BLUE = Font(name=FONT, size=10, color="0000FF")
CUR = '#,##0;(#,##0);-'; PCT = '0.00%'
CEN = Alignment(horizontal="center"); RGT = Alignment(horizontal="right")


def _s(ws, r, c, v, font=REG, fmt=None, fill=None, al=None):
    cell = ws.cell(row=r, column=c, value=v); cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if al: cell.alignment = al
    return cell


def _income_sheet(wb, ex, cfg, idx):
    """גיליון הכנסות: מדרגה אלכסונית עם נוסחאות. מחזיר (שם גיליון, שורת סה\"כ, עמודת חודש ראשון)."""
    name = f"תזרים הכנסות מתחם {idx}"
    ws = wb.create_sheet(name); ws.sheet_view.rightToLeft = True
    M = cfg["months"]
    _s(ws, 1, 2, f"תזרים הכנסות — מתחם {idx}", Font(name=FONT, bold=True, size=13, color="4F2D7F"))
    # פרמטרים (קלט)
    use20 = 1 if cfg.get("use_track20", True) else 0
    _s(ws, 3, 2, "פדיון (ללא מע\"מ)", REG, al=RGT); _s(ws, 3, 3, round(ex["pidyon"]), BLUE, fmt=CUR, fill=INP)
    _s(ws, 4, 2, "חודשי תזרים", REG, al=RGT); _s(ws, 4, 3, M, BLUE, fill=INP)
    _s(ws, 5, 2, "תקבול ראשון (מקדמה)", REG, al=RGT); _s(ws, 5, 3, cfg["down_payment"], BLUE, fmt=PCT, fill=INP)
    _s(ws, 6, 2, "מסלול 20/80 פעיל", REG, al=RGT); _s(ws, 6, 3, use20, BLUE, fill=INP)
    _s(ws, 7, 2, "חלק מסלול 20%", REG, al=RGT); _s(ws, 7, 3, cfg["track20_share"], BLUE, fmt=PCT, fill=INP)
    P_PID, P_M, P_DOWN, P_U20, P_SH = "$C$3", "$C$4", "$C$5", "$C$6", "$C$7"
    _s(ws, 8, 2, "פדיון מסלול 20%", REG, al=RGT); _s(ws, 8, 3, f"={P_PID}*{P_U20}*{P_SH}", REG, fmt=CUR); I4 = "$C$8"
    _s(ws, 9, 2, "פדיון מסלול רגיל", REG, al=RGT); _s(ws, 9, 3, f"={P_PID}-{I4}", REG, fmt=CUR); I5 = "$C$9"
    _s(ws, 10, 2, "אצווה חודשית (רגיל)", REG, al=RGT); _s(ws, 10, 3, f"={I5}/{P_M}", REG, fmt=CUR)
    BATCH = "$C$10"

    MC0 = 4
    def mcol(m): return MC0 + m
    hr = 12; midx = 13
    _s(ws, hr, 2, "אצווה \\ חודש", WHT, fill=HDR)
    _s(ws, midx, 2, "אינדקס", REG, al=RGT)
    for m in range(1, M + 1):
        _s(ws, hr, mcol(m), m, WHT, fill=HDR, al=CEN)
        _s(ws, midx, mcol(m), m, REG, al=CEN)
    tot_col = mcol(M) + 1
    _s(ws, hr, tot_col, "סה\"כ", WHT, fill=HDR)

    b0 = midx + 1
    # שורת מסלול 20%: מקדמה בחודש 1, בלון במסירה
    t20 = b0
    _s(ws, t20, 2, "מסלול 20% (חתימה+בלון)", REG, al=RGT)
    for m in range(1, M + 1):
        col = mcol(m)
        f = f"=IF({CL(col)}${midx}=1,{I4}*{P_DOWN},IF({CL(col)}${midx}={P_M},{I4}*(1-{P_DOWN}),0))"
        _s(ws, t20, col, f, REG, fmt=CUR)
    _s(ws, t20, tot_col, f"=SUM({CL(mcol(1))}{t20}:{CL(mcol(M))}{t20})", REG, fmt=CUR)
    # אצוות רגילות
    reg0 = t20 + 1
    for k in range(1, M + 1):
        rr = reg0 + (k - 1)
        _s(ws, rr, 2, f"אצווה חודש {k}", REG, al=RGT)
        for m in range(1, M + 1):
            col = mcol(m)
            if m < k: f = 0
            elif m == k: f = f"={BATCH}*{P_DOWN}"
            else: f = f"=IF({P_M}-{k}>0,{BATCH}*(1-{P_DOWN})/({P_M}-{k}),0)"
            _s(ws, rr, col, f, REG, fmt=CUR)
        if k == M: _s(ws, rr, mcol(k), f"={BATCH}", REG, fmt=CUR)
        _s(ws, rr, tot_col, f"=SUM({CL(mcol(1))}{rr}:{CL(mcol(M))}{rr})", REG, fmt=CUR)
    last = reg0 + M - 1
    tr = last + 1
    _s(ws, tr, 2, "הכנסות לחודש", BLD, fill=SUB)
    for m in range(1, M + 1):
        col = mcol(m)
        _s(ws, tr, col, f"=SUM({CL(col)}{t20}:{CL(col)}{last})", BLD, fmt=CUR, fill=SUB)
    _s(ws, tr, tot_col, f"=SUM({CL(mcol(1))}{tr}:{CL(mcol(M))}{tr})", BLD, fmt=CUR, fill=SUB)

    ws.column_dimensions["B"].width = 18
    for m in range(M + 1): ws.column_dimensions[CL(MC0 + m)].width = 12
    return name, tr, MC0


SPREAD = {  # כלל -> נוסחה לתא חודשי m (T=תא סכום, mi=תא אינדקס, P=מילון תאי הנחות)
    "at_permit": lambda T, mi, P: f"=IF({mi}=0,{T},0)",
    "month1": lambda T, mi, P: f"=IF({mi}=1,{T},0)",
    "linear": lambda T, mi, P: f"=IF({mi}>=1,{T}/{P['months']},0)",
    "permit_linear": lambda T, mi, P: f"=IF({mi}=0,{T}*{P['permit']},IF({mi}>=1,{T}*(1-{P['permit']})/{P['months']},0))",
    "start_end": lambda T, mi, P: f"=IF({mi}=1,{T}/2,IF({mi}={P['months']},{T}/2,0))",
    "rent": lambda T, mi, P: (f"=IF(AND({mi}>=1,MOD({mi}-1,12)=0),MIN({T}/{P['months']}*12,"
                              f"MAX({T}-({T}/{P['months']}*12)*INT(({mi}-1)/12),0)),0)"),
    "win_bsmt": lambda T, mi, P: f"=IF(AND({mi}>={P['bs']},{mi}<={P['bs']}+{P['bd']}-1),{T}/{P['bd']},0)",
    "win_skel": lambda T, mi, P: f"=IF(AND({mi}>={P['ss']},{mi}<={P['ss']}+{P['sd']}-1),{T}/{P['sd']},0)",
    "win_fin": lambda T, mi, P: f"=IF(AND({mi}>={P['fs']},{mi}<={P['fs']}+{P['fd']}-1),{T}/{P['fd']},0)",
}


def _expense_sheet(wb, ex, cfg, idx, inc_name, inc_row, inc_mc0):
    name = f"תזרים הוצאות מתחם {idx}"
    ws = wb.create_sheet(name); ws.sheet_view.rightToLeft = True
    M = cfg["months"]; n = M + 1
    _s(ws, 1, 2, f"תזרים הוצאות ומימון — מתחם {idx}", Font(name=FONT, bold=True, size=13, color="4F2D7F"))
    # הנחות עבודה (קלט)
    assum = [("rate", "ריבית שנתית", cfg["annual_rate"], PCT), ("equity", "הון עצמי", cfg["equity_pct"], PCT),
             ("permit", "תכנון בהיתר", cfg["planning_permit_share"], PCT),
             ("fee_acc", "עמלת ליווי", cfg["fee_accompaniment"], PCT), ("fee_sale", "ערבות חוק מכר", cfg["fee_sale_law"], PCT),
             ("fee_rent", "ערבות שכ\"ד", cfg["fee_rent"], PCT), ("fee_own", "ערבות בעלים", cfg["fee_owners"], PCT),
             ("fee_nu", "עמלת אי-ניצול", cfg["fee_non_util"], PCT), ("months", "חודשי בנייה", M, "0"),
             ("bs", "מרתף התחלה", cfg["basement"][0], "0"), ("bd", "מרתף משך", cfg["basement"][1], "0"),
             ("ss", "שלד התחלה", cfg["skeleton"][0], "0"), ("sd", "שלד משך", cfg["skeleton"][1], "0"),
             ("fs", "גמר התחלה", cfg["finishing"][0], "0"), ("fd", "גמר משך", cfg["finishing"][1], "0")]
    P = {}; ar = 3
    for i, (k, lbl, val, fmt) in enumerate(assum):
        _s(ws, ar + i, 2, lbl, REG, al=RGT); _s(ws, ar + i, 3, val, BLUE, fmt=fmt, fill=INP)
        P[k] = f"$C${ar+i}"
    # ערכי בסיס
    br = ar + len(assum) + 1
    _s(ws, br, 2, "פדיון (לערבות בעלים)", REG, al=RGT); _s(ws, br, 3, round(ex["pidyon"]), REG, fmt=CUR); P["pidyon"] = f"$C${br}"
    _s(ws, br+1, 2, "סה\"כ עלות (לפני מימון)", REG, al=RGT); _s(ws, br+1, 3, round(ex["total_before_fin"]), REG, fmt=CUR); P["cost"] = f"$C${br+1}"
    _s(ws, br+2, 2, "שווי דירות דיירים", REG, al=RGT); _s(ws, br+2, 3, round(ex["owners"]), REG, fmt=CUR); P["owners"] = f"$C${br+2}"

    hr = br + 3; mir = hr + 1
    MC0 = 4
    def mcol(m): return MC0 + m
    def mi(m): return f"{CL(mcol(m))}${mir}"
    tot_col = mcol(M) + 1
    _s(ws, hr, 2, "סעיף \\ חודש", WHT, fill=HDR); _s(ws, mir, 2, "אינדקס", REG, al=RGT)
    for m in range(n):
        _s(ws, hr, mcol(m), "היתר" if m == 0 else m, WHT, fill=HDR, al=CEN)
        _s(ws, mir, mcol(m), m, REG, al=CEN)
    _s(ws, hr, tot_col, "סה\"כ", WHT, fill=HDR)

    # שורות עלות: סכום (ערך) + פריסה בנוסחה
    r = mir + 1; rent_rows = []
    for nm, amt, rule in ex["cost_lines"]:
        _s(ws, r, 2, nm[:32], REG, al=RGT)
        _s(ws, r, 3, round(amt), REG, fmt=CUR)      # סכום מחולץ (ערך)
        T = f"$C${r}"
        for m in range(n):
            _s(ws, r, mcol(m), SPREAD.get(rule, SPREAD["linear"])(T, mi(m), P), REG, fmt=CUR)
        _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR)
        if rule == "rent": rent_rows.append(f"$C${r}")
        r += 1
    exp_r = r
    _s(ws, r, 2, "סה\"כ הוצאות", BLD, fill=SUB)
    for m in range(n):
        _s(ws, r, mcol(m), f"=SUM({CL(mcol(m))}{mir+1}:{CL(mcol(m))}{exp_r-1})", BLD, fmt=CUR, fill=SUB)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{exp_r}:{CL(mcol(M))}{exp_r})", BLD, fmt=CUR, fill=SUB)
    EXP_T = f"${CL(tot_col)}${exp_r}"; r += 1
    rent_total = "(" + "+".join(rent_rows) + ")" if rent_rows else "0"

    # הכנסות (מגיליון ההכנסות)
    inc_r = r; _s(ws, r, 2, "הכנסות", REG)
    for m in range(n):
        if m == 0: _s(ws, r, mcol(m), 0, REG, fmt=CUR)
        else: _s(ws, r, mcol(m), f"='{inc_name}'!{CL(inc_mc0+m)}{inc_row}", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(1))}{inc_r}:{CL(mcol(M))}{inc_r})", REG, fmt=CUR); r += 1

    # ערבויות
    g0 = r
    _s(ws, r, 2, "עמלת ליווי", REG, al=RGT)
    for m in range(n): _s(ws, r, mcol(m), f"=IF({mi(m)}=1,{EXP_T}*{P['fee_acc']},0)", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR); acc_r = r; r += 1
    _s(ws, r, 2, "ערבות חוק מכר", REG, al=RGT)   # על הכנסות החודש
    for m in range(n):
        if m == 0: _s(ws, r, mcol(m), 0, REG, fmt=CUR)
        else: _s(ws, r, mcol(m), f"={CL(mcol(m))}{inc_r}*{P['fee_sale']}/12", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR); sale_r = r; r += 1
    _s(ws, r, 2, "ערבות שכ\"ד", REG, al=RGT)
    for m in range(n): _s(ws, r, mcol(m), f"=IF({mi(m)}>=1,{P['fee_rent']}/12*({rent_total}/{P['months']}*12),0)", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR); rent_r = r; r += 1
    _s(ws, r, 2, "ערבות בעלים", REG, al=RGT)     # על שווי דירות הדיירים
    for m in range(n): _s(ws, r, mcol(m), f"=IF({mi(m)}>=1,{P['fee_own']}/12*{P['owners']},0)", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR); own_r = r; r += 1
    _s(ws, r, 2, "עמלת אי-ניצול", REG, al=RGT)
    for m in range(n):
        cum = f"SUM({CL(mcol(0))}{exp_r}:{CL(mcol(m))}{exp_r})"
        _s(ws, r, mcol(m), f"=IF({mi(m)}>=1,{P['fee_nu']}/12*MAX({EXP_T}-{cum},0),0)", REG, fmt=CUR)
    _s(ws, r, tot_col, f"=SUM({CL(mcol(0))}{r}:{CL(mcol(M))}{r})", REG, fmt=CUR); nu_r = r; g1 = r; r += 2

    # בלוק מימון
    _s(ws, r, 2, "סך הוצאות כולל עמלות", BLD); incl_r = r
    for m in range(n):
        _s(ws, r, mcol(m), f"={CL(mcol(m))}{exp_r}+SUM({CL(mcol(m))}{g0}:{CL(mcol(m))}{g1})", BLD, fmt=CUR)
    r += 1
    _s(ws, r, 2, "תזרים חודשי", REG); cf_r = r
    for m in range(n): _s(ws, r, mcol(m), f"={CL(mcol(m))}{inc_r}-{CL(mcol(m))}{incl_r}", REG, fmt=CUR)
    r += 1
    _s(ws, r, 2, "הון עצמי", REG); eq_r = r; cap = f"{P['equity']}*{EXP_T}"
    for m in range(n):
        if m == 0: f = f"=MAX(MIN({CL(mcol(m))}{incl_r},{cap}),0)"
        else:
            prior = f"SUM({CL(mcol(0))}{eq_r}:{CL(mcol(m-1))}{eq_r})"
            f = f"=MAX(MIN({CL(mcol(m))}{incl_r},{cap}-{prior}),0)"
        _s(ws, r, mcol(m), f, REG, fmt=CUR)
    r += 1
    _s(ws, r, 2, "יתרה לפני ריבית", REG); bb_r = r
    for m in range(n):
        if m == 0: f = f"={CL(mcol(m))}{eq_r}+{CL(mcol(m))}{cf_r}"
        else: f = f"={CL(mcol(m-1))}{bb_r+2}+{CL(mcol(m))}{eq_r}+{CL(mcol(m))}{cf_r}"
        _s(ws, r, mcol(m), f, REG, fmt=CUR)
    r += 1
    _s(ws, r, 2, "ריבית", REG); int_r = r
    for m in range(n): _s(ws, r, mcol(m), f"=IF({CL(mcol(m))}{bb_r}<0,{CL(mcol(m))}{bb_r}*{P['rate']}/12,0)", REG, fmt=CUR)
    r += 1
    _s(ws, r, 2, "יתרה לסוף חודש", BLD); ai_r = r
    for m in range(n): _s(ws, r, mcol(m), f"={CL(mcol(m))}{bb_r}+{CL(mcol(m))}{int_r}", BLD, fmt=CUR)
    r += 2

    # סיכום מימון
    _s(ws, r, 2, "סיכום מימון", Font(name=FONT, bold=True, size=12, color="4F2D7F")); r += 1
    fr0 = r
    for lbl, ref in [("ריבית נצברת", f"=-SUM({CL(mcol(0))}{int_r}:{CL(mcol(M))}{int_r})"),
                     ("עמלת ליווי", f"={CL(tot_col)}{acc_r}"), ("ערבות חוק מכר", f"={CL(tot_col)}{sale_r}"),
                     ("ערבות שכ\"ד", f"={CL(tot_col)}{rent_r}"), ("ערבות בעלים", f"={CL(tot_col)}{own_r}"),
                     ("עמלת אי-ניצול", f"={CL(tot_col)}{nu_r}")]:
        _s(ws, r, 2, lbl, REG, al=RGT); _s(ws, r, 3, ref, REG, fmt=CUR, fill=FIN); r += 1
    _s(ws, r, 2, "סה\"כ מימון", BLD, al=RGT); _s(ws, r, 3, f"=ROUND(SUM(C{fr0}:C{r-1}),-4)", BLD, fmt=CUR, fill=TOT); tf_r = r; r += 1
    _s(ws, r, 2, "אחוז מימון", Font(name=FONT, bold=True, size=12), al=RGT)
    _s(ws, r, 3, f"=C{tf_r}/{P['cost']}", Font(name=FONT, bold=True, size=12, color="4F2D7F"), fmt=PCT, fill=TOT)
    pct_cell = f"'{name}'!$C${r}"; totfin_cell = f"'{name}'!$C${tf_r}"

    ws.column_dimensions["B"].width = 26
    for m in range(n + 1): ws.column_dimensions[CL(MC0 + m)].width = 12
    ws.freeze_panes = f"{CL(MC0)}{mir+1}"
    return pct_cell, totfin_cell


def build(path_in, path_out, overrides=None):
    from generate_cashflow import CFG
    cfg = dict(CFG)
    if overrides: cfg.update({k: v for k, v in overrides.items() if v is not None})
    wb = Workbook(); wb.remove(wb.active)
    summary = wb.create_sheet("סיכום"); summary.sheet_view.rightToLeft = True
    _s(summary, 1, 2, "סיכום מימון לפי מתחם", Font(name=FONT, bold=True, size=14, color="4F2D7F"))
    for j, h in enumerate(["מתחם", "פדיון", "עלות לפני מימון", "סה\"כ מימון", "אחוז מימון", "בקרה"]):
        _s(summary, 3, 2 + j, h, WHT, fill=HDR, al=CEN)
    from finance_multi import finance_compound
    results = []; row = 4
    for i, sh in enumerate(X.find_compound_sheets(path_in), start=1):
        ex = X.extract_compound(path_in, sh)
        c2 = dict(cfg)
        if ex["build_months"]:
            c2["months"] = int(ex["build_months"])
        inc_name, inc_row, inc_mc0 = _income_sheet(wb, ex, c2, i)
        pct_cell, totfin_cell = _expense_sheet(wb, ex, c2, i, inc_name, inc_row, inc_mc0)
        fc = finance_compound(ex, c2)      # לתצוגת המסך (תואם לנוסחאות)
        ok = ex["reconciled"]
        _s(summary, row, 2, sh, REG, al=RGT)
        _s(summary, row, 3, round(ex["pidyon"]), REG, fmt=CUR)
        _s(summary, row, 4, round(ex["total_before_fin"]), REG, fmt=CUR)
        _s(summary, row, 5, f"={totfin_cell}", REG, fmt=CUR)
        _s(summary, row, 6, f"={pct_cell}", BLD, fmt=PCT)
        _s(summary, row, 7, "תקין ✓" if ok else "פער — בדוק", REG, fill=OKF if ok else BADF, al=CEN)
        results.append(dict(sheet=sh, pidyon=ex["pidyon"], owners=ex["owners"],
                            cost=ex["total_before_fin"], financing=fc["breakdown"].total,
                            pct=fc["pct"], reconciled=ok))
        row += 1
    summary.column_dimensions["B"].width = 20
    for c in range(3, 8): summary.column_dimensions[CL(c)].width = 18
    wb.save(path_out)
    return results


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "out_formulas.xlsx")
    print("נוצר.")
