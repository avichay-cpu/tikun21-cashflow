"""
חילוץ נתוני מקור לפי תוויות וחלקים — עמיד לשינויי מבנה
======================================================
קורא גיליונות "תחשיב מתחם N", מזהה את חמשת חלקי העלות לפי הכותרות
הממוספרות, ומחלץ שורות עלות + פדיון (J) + שווי דירות דיירים (K).
מיפוי לפי תוויות — לא לפי מספרי שורה — כדי לעבוד על מבנים שונים.
"""
from openpyxl import load_workbook


def _num(x):
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(",", "").replace("₪", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _label(ws, r):
    for col in ("B", "C", "D"):
        v = ws[f"{col}{r}"].value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def find_compound_sheets(path):
    """כל הגיליונות ששמם 'תחשיב מתחם N' (ממויין לפי מספר)."""
    wb = load_workbook(path, read_only=True)
    names = [s for s in wb.sheetnames if s.strip().startswith("תחשיב מתחם")]
    wb.close()
    def num(nm):
        digits = "".join(ch for ch in nm if ch.isdigit())
        return int(digits) if digits else 0
    return sorted(names, key=num)


# כללי פריסה לפי חלק (section index 1..5)
SECTION_RULE = {1: "at_permit", 3: "linear", 5: "at_permit"}


def extract_compound(path, sheet):
    """מחלץ נתוני מתחם בודד: פדיון, דיירים, שורות עלות עם כללי פריסה, ואימות."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    sections = []          # [{idx,name,lines:[(label,val)],subtotal}]
    cur = None
    pidyon = owners = 0.0
    units = build_months = 0.0
    stated_total = None

    for r in range(1, min(ws.max_row, 320) + 1):
        lbl = _label(ws, r)
        if not lbl:
            continue
        # כותרת חלק ממוספרת
        if lbl[:2] in ("1.", "2.", "3.", "4.", "5.", "6.") and any(
                k in lbl for k in ("הוצאות", "עלות", "מיסוי")):
            cur = {"idx": int(lbl[0]), "name": lbl, "lines": [], "subtotal": None}
            sections.append(cur)
            continue
        # פדיון + דירות דיירים (השורה מתחילה ב"סה\"כ", לכן נבדק לפני דילוג הסיכומים)
        if "פדיון חזוי ליזם" in lbl and "ללא מע" in lbl:
            # זיהוי עמודת הפדיון לפי כותרת "פדיון ללא מע\"מ" (העמודה השמאלית מבין ה'ללא מע"מ')
            pcol = None
            for hr in range(max(1, r - 15), r):
                cols = [c for c in range(7, 15)
                        if isinstance(ws.cell(row=hr, column=c).value, str)
                        and "ללא מע" in ws.cell(row=hr, column=c).value]
                if cols:
                    pcol = min(cols); break
            pidyon = _num(ws.cell(row=r, column=pcol).value) if pcol else _num(ws[f"J{r}"].value)
            owners = _num(ws[f"K{r}"].value)      # שווי דירות דיירים (גיבוי בלבד)
            continue
        if "דירות לשיווק" in lbl and not units:
            units = _num(ws[f"C{r}"].value)       # מספר דירות לשיווק
            continue
        if "תקופה מוערכת לבני" in lbl and not build_months:
            build_months = _num(ws[f"C{r}"].value)  # חודשי בנייה מהקובץ
        # סה"כ עלות הקמה (לפני מימון = הנמוך) — מסמן סוף פירוט העלויות
        if "עלות הקמת הפרוי" in lbl:
            v = _num(ws[f"G{r}"].value)
            if v:
                stated_total = v if stated_total is None else min(stated_total, v)
            cur = None
            continue
        # שורת סיכום חלק — ודילוג על כל שורת "סה\"כ" (ביניים) מהאיסוף
        if lbl.startswith("סה"):
            if cur is not None and cur["subtotal"] is None:
                v = _num(ws[f"G{r}"].value)
                if v:
                    cur["subtotal"] = v
            continue
        # שורת עלות רגילה בתוך חלק (לא כולל שורות מימון/תזרים)
        if cur is not None and "תזרים" not in lbl and "מימון" not in lbl:
            v = _num(ws[f"G{r}"].value)
            if v:
                cur["lines"].append((lbl, v))
    wb.close()

    sec_sum = sum(s["subtotal"] or 0 for s in sections)
    total_before_fin = stated_total if stated_total else sec_sum

    # שווי דירות הדיירים לערבות בעלים = מחיר למ"ר × שטח דירות (עיקרי + מרפסות×0.5)
    owners_value = _owners_value(ws)

    cost_lines = _build_cost_lines(sections)
    computed = sum(a for _, a, _ in cost_lines)

    return dict(
        sheet=sheet, pidyon=pidyon, owners=owners_value or owners, units=units, build_months=build_months,
        total_before_fin=total_before_fin, sec_sum=sec_sum,
        computed_cost=computed, cost_lines=cost_lines,
        reconciled=abs(computed - total_before_fin) < max(1000, total_before_fin * 0.005),
    )


def _owners_value(ws):
    """שווי דירות הדיירים = מחיר למ\"ר × שטח (עיקרי + מרפסות×0.5).
    מחיר: שורת "שווי מ\"ר ממוצע לשיווק". שטח: שורת "סה\"כ" שאחרי "מיוחדות"."""
    price = 0.0
    for r in range(1, min(ws.max_row, 300) + 1):
        for c in range(1, min(ws.max_column, 16) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "שווי" in v and "שיווק" in v and "יח" not in v:
                for cc in range(c + 1, min(ws.max_column, 16) + 1):
                    nv = ws.cell(row=r, column=cc).value
                    if isinstance(nv, (int, float)) and nv > 1000:
                        price = float(nv); break
                if price:
                    break
        if price:
            break
    if not price:
        return 0.0
    # שטח דירות: שורת "סה\"כ" שמיד אחרי שורת "מיוחדות"
    special_r = None
    for r in range(1, min(ws.max_row, 300) + 1):
        if "מיוחדות" in _label(ws, r):
            special_r = r; break
    area_r = None
    if special_r:
        for r in range(special_r + 1, special_r + 6):
            if _label(ws, r).startswith("סה"):
                area_r = r; break
    if not area_r:
        return 0.0
    main = _num(ws[f"I{area_r}"].value)
    balc = _num(ws[f"J{area_r}"].value) + _num(ws[f"K{area_r}"].value)
    return price * main + price * balc * 0.5


def _build_cost_lines(sections):
    """ממיר שורות חלקים לשורות עלות עם כללי פריסה."""
    lines = []
    for s in sections:
        idx = s["idx"]
        if idx == 2:
            # בנייה ישירה: מרתף (תת-קרקעי) + שלד 60% + גמר 40%
            total = s["subtotal"] or sum(v for _, v in s["lines"])
            basement = 0.0
            for lbl, v in s["lines"]:
                if "תת קרקעי" in lbl or "תת-קרקעי" in lbl:
                    basement += v
            above = total - basement
            lines.append(("מרתפים", basement, "win_bsmt"))
            lines.append(("שלד", above * 0.6, "win_skel"))
            lines.append(("גמר", above * 0.4, "win_fin"))
            continue
        for lbl, v in s["lines"]:
            rule = SECTION_RULE.get(idx, "linear")
            if idx == 3 and "תכנון" in lbl:
                rule = "permit_linear"
            elif idx == 4:
                if "שכ" in lbl and "דיור חלוף" in lbl:
                    rule = "rent"
                elif "העברה" in lbl or "הובל" in lbl:
                    rule = "start_end"
                else:
                    rule = "linear"
            lines.append((lbl, v, rule))
    return lines


if __name__ == "__main__":
    import sys
    path = sys.argv[1]
    for sh in find_compound_sheets(path):
        e = extract_compound(path, sh)
        flag = "✓ מאוזן" if e["reconciled"] else "✗ פער"
        print(f"[{sh}] פדיון={e['pidyon']:,.0f} דיירים={e['owners']:,.0f} "
              f"עלות={e['total_before_fin']:,.0f} חושב={e['computed_cost']:,.0f} {flag} "
              f"({len(e['cost_lines'])} שורות)")
