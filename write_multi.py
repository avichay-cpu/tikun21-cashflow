"""
כתיבת קובץ תזרים רב-מתחמי
=========================
לכל מתחם: גיליון "תזרים הוצאות מתחם N" + "תזרים הכנסות מתחם N".
גיליון ראשון "סיכום" עם אחוז מימון לכל מתחם ובקרת ביטחון.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

import extract as X
from finance_multi import finance_compound, RULEMAP
from generate_cashflow import build_revenue, CFG

FONT = "Arial"
HDR = PatternFill("solid", fgColor="4F2D7F"); SUB = PatternFill("solid", fgColor="EDE7F6")
TOT = PatternFill("solid", fgColor="FFE0B2"); FIN = PatternFill("solid", fgColor="FFF3E0")
OK = PatternFill("solid", fgColor="C8E6C9"); BAD = PatternFill("solid", fgColor="FFCDD2")
WHT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
BLD = Font(name=FONT, bold=True, size=10); REG = Font(name=FONT, size=10)
CUR = '#,##0;(#,##0);-'; PCT = '0.00%'
CEN = Alignment(horizontal="center"); RGT = Alignment(horizontal="right")


def _set(ws, r, c, v, font=REG, fmt=None, fill=None, al=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if al: cell.alignment = al
    return cell


def _write_compound(wb, ex, cfg, idx):
    """כותב שני גיליונות למתחם: הוצאות (עם פריסה) והכנסות."""
    n = cfg["months"] + 1
    fc = finance_compound(ex, cfg)
    # ---- תזרים הוצאות ----
    ws = wb.create_sheet(f"תזרים הוצאות מתחם {idx}")
    ws.sheet_view.rightToLeft = True
    _set(ws, 1, 2, f"תזרים הוצאות ומימון — מתחם {idx}", Font(name=FONT, bold=True, size=13, color="4F2D7F"))
    MC0 = 4
    def mcol(m): return MC0 + m
    tot_col = mcol(cfg["months"]) + 1
    hr = 3
    _set(ws, hr, 2, "סעיף / חודש", WHT, fill=HDR)
    for m in range(n):
        _set(ws, hr, mcol(m), "היתר" if m == 0 else m, WHT, fill=HDR, al=CEN)
    _set(ws, hr, tot_col, "סה\"כ", WHT, fill=HDR)

    # שורות עלות עם פריסה (ערכים)
    r = hr + 1
    base_expense = [0.0] * n
    for name, amt, rule in ex["cost_lines"]:
        vec = RULEMAP.get(rule, RULEMAP["linear"])(amt, n, cfg)
        _set(ws, r, 2, name[:34], REG, al=RGT)
        for m in range(n):
            _set(ws, r, mcol(m), round(vec[m]), REG, fmt=CUR)
            base_expense[m] += vec[m]
        _set(ws, r, tot_col, f"=SUM({CL(MC0)}{r}:{CL(mcol(cfg['months']))}{r})", REG, fmt=CUR)
        r += 1
    exp_r = r
    _set(ws, r, 2, "סה\"כ הוצאות", BLD, fill=SUB)
    for m in range(n):
        _set(ws, r, mcol(m), f"=SUM({CL(mcol(m))}{hr+1}:{CL(mcol(m))}{exp_r-1})", BLD, fmt=CUR, fill=SUB)
    _set(ws, r, tot_col, f"=SUM({CL(MC0)}{exp_r}:{CL(mcol(cfg['months']))}{exp_r})", BLD, fmt=CUR, fill=SUB)
    r += 2

    # בלוק מימון (ערכים מהמנוע)
    revenue = build_revenue(ex["pidyon"], cfg)
    fb = fc["breakdown"]
    # נבנה מחדש את הזרמים לצורך הצגה
    from finance_multi import finance_compound as _fc
    total_expense_base = sum(base_expense)
    acc = total_expense_base * cfg["fee_accompaniment"]
    sale = [revenue[m] * cfg["fee_sale_law"] / 12 for m in range(n)]
    rent_total = sum(a for nm, a, rl in ex["cost_lines"] if rl == "rent")
    rent_g = cfg["fee_rent"] / 12 * (rent_total / cfg["months"] * 12) if rent_total else 0.0
    own_full = cfg["fee_owners"] / 12 * ex["owners"]
    non_util = [0.0] * n; spent = base_expense[0]
    for m in range(1, n):
        spent += base_expense[m]; non_util[m] = max(total_expense_base - spent, 0.0) * cfg["fee_non_util"] / 12
    guar = [0.0]*n; guar[1] += acc
    for m in range(1, n): guar[m] += sale[m] + rent_g + own_full + non_util[m]
    incl = [base_expense[m] + guar[m] for m in range(n)]
    from financing import simulate_financing
    fr = simulate_financing(revenue, incl, total_expense_base*cfg["equity_pct"], cfg["annual_rate"])

    for label, series in [("ערבויות ועמלות", guar), ("סך הוצאות כולל עמלות", incl),
                          ("הכנסות (פדיון)", revenue),
                          ("תזרים חודשי", [revenue[m]-incl[m] for m in range(n)]),
                          ("הון עצמי", fr.equity_injection),
                          ("יתרה לפני ריבית", fr.balance_before_interest),
                          ("ריבית", fr.interest), ("יתרה לסוף חודש", fr.balance)]:
        bold = label in ("סך הוצאות כולל עמלות", "יתרה לסוף חודש")
        _set(ws, r, 2, label, BLD if bold else REG, al=RGT)
        for m in range(n):
            _set(ws, r, mcol(m), round(series[m]), BLD if bold else REG, fmt=CUR)
        _set(ws, r, tot_col, round(sum(series)), BLD if bold else REG, fmt=CUR)
        r += 1
    r += 1

    # סיכום מימון
    _set(ws, r, 2, "סיכום מימון", Font(name=FONT, bold=True, size=12, color="4F2D7F")); r += 1
    for lbl, val in [("ריבית נצברת", fb.interest), ("עמלת ליווי", fb.accompaniment),
                     ("ערבות חוק מכר", fb.sale_law_guarantee), ("ערבות שכ\"ד", fb.rent_guarantee),
                     ("ערבות בעלים", fb.owners_guarantee), ("עמלת אי-ניצול", fb.non_utilization)]:
        _set(ws, r, 2, lbl, REG, al=RGT); _set(ws, r, 3, round(val), REG, fmt=CUR, fill=FIN); r += 1
    _set(ws, r, 2, "סה\"כ מימון", BLD, al=RGT); _set(ws, r, 3, fb.total, BLD, fmt=CUR, fill=TOT); r += 1
    _set(ws, r, 2, "סה\"כ עלות (לפני מימון)", REG, al=RGT); _set(ws, r, 3, round(ex["total_before_fin"]), REG, fmt=CUR); r += 1
    _set(ws, r, 2, "אחוז מימון", Font(name=FONT, bold=True, size=12), al=RGT)
    _set(ws, r, 3, fc["pct"], Font(name=FONT, bold=True, size=12, color="4F2D7F"), fmt=PCT, fill=TOT)

    ws.column_dimensions["B"].width = 28
    for m in range(n + 1): ws.column_dimensions[CL(MC0 + m)].width = 12
    ws.freeze_panes = f"{CL(MC0)}{hr+1}"
    return fc


def build(path_in, path_out, overrides=None):
    cfg = dict(CFG)
    if overrides:
        cfg.update({k: v for k, v in overrides.items() if v is not None})
    wb = Workbook(); wb.remove(wb.active)
    summary = wb.create_sheet("סיכום"); summary.sheet_view.rightToLeft = True
    _set(summary, 1, 2, "סיכום מימון לפי מתחם", Font(name=FONT, bold=True, size=14, color="4F2D7F"))
    hdr = ["מתחם", "פדיון (ללא מע\"מ)", "עלות לפני מימון", "סה\"כ מימון", "אחוז מימון", "בקרת ביטחון"]
    for j, h in enumerate(hdr):
        _set(summary, 3, 2 + j, h, WHT, fill=HDR, al=CEN)

    results = []
    row = 4
    for i, sh in enumerate(X.find_compound_sheets(path_in), start=1):
        ex = X.extract_compound(path_in, sh)
        fc = _write_compound(wb, ex, cfg, i)
        ok = ex["reconciled"]
        _set(summary, row, 2, sh, REG, al=RGT)
        _set(summary, row, 3, round(ex["pidyon"]), REG, fmt=CUR)
        _set(summary, row, 4, round(ex["total_before_fin"]), REG, fmt=CUR)
        _set(summary, row, 5, round(fc["breakdown"].total), REG, fmt=CUR)
        _set(summary, row, 6, fc["pct"], BLD, fmt=PCT)
        _set(summary, row, 7, "תקין ✓" if ok else "פער — בדוק", REG, fill=OK if ok else BAD, al=CEN)
        results.append(dict(sheet=sh, pidyon=ex["pidyon"], owners=ex["owners"],
                            cost=ex["total_before_fin"], financing=fc["breakdown"].total,
                            pct=fc["pct"], reconciled=ok))
        row += 1
    summary.column_dimensions["B"].width = 20
    for c in range(3, 8): summary.column_dimensions[CL(c)].width = 18
    wb.save(path_out)
    return results


if __name__ == "__main__":
    res = build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "out_multi.xlsx")
    for r in res:
        print(f"{r['sheet']}: אחוז={r['pct']:.2%} {'✓' if r['reconciled'] else '✗'}")
